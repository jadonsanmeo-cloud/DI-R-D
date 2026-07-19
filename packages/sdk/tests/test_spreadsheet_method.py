from __future__ import annotations

from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from data_intelligence_sdk.methods.spreadsheet import (
    SpreadsheetReadError,
    detect_spreadsheet_format,
    materialize_spreadsheet,
    register_spreadsheet_methods,
)
from data_intelligence_sdk.runtime.method_hub import MethodHub


class SpreadsheetMethodTests(unittest.TestCase):
    def test_detects_legacy_xls_from_container_signature(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "legacy.xlsx"
            path.write_bytes(bytes.fromhex("D0CF11E0A1B11AE1") + b"payload")

            self.assertEqual(detect_spreadsheet_format(path), "xls")

    def test_rejects_unknown_container_instead_of_returning_no_data(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "invalid.xls"
            path.write_bytes(b"not an excel workbook")

            with self.assertRaisesRegex(
                SpreadsheetReadError, "Unsupported spreadsheet container"
            ):
                materialize_spreadsheet(str(path))

    def test_materializes_xlsx_with_header_detection_and_json_values(self):
        from openpyxl import Workbook

        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "scores.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Scores"
            sheet.append(["Student score report"])
            sheet.append(["Student", "Score", "Score", None])
            sheet.append(["An", 8.5, 9, datetime(2026, 7, 19, 10, 30)])
            sheet.append(["Binh", 7, 8, None])
            workbook.save(path)
            workbook.close()

            rows = materialize_spreadsheet(str(path))

        self.assertEqual(len(rows), 2)
        self.assertEqual(
            list(rows[0]),
            ["Student", "Score", "Score_2", "column_4"],
        )
        self.assertEqual(rows[0]["Student"], "An")
        self.assertEqual(rows[0]["Score"], 8.5)
        self.assertEqual(rows[0]["column_4"], "2026-07-19T10:30:00")

    def test_can_combine_non_empty_sheets_with_lineage(self):
        from openpyxl import Workbook

        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "multi.xlsx"
            workbook = Workbook()
            first = workbook.active
            first.title = "First"
            first.append(["Name", "Value"])
            first.append(["A", 1])
            second = workbook.create_sheet("Second")
            second.append(["Name", "Value"])
            second.append(["B", 2])
            workbook.save(path)
            workbook.close()

            rows = materialize_spreadsheet(str(path), include_all_sheets=True)

        self.assertEqual([row["__sheet_name"] for row in rows], ["First", "Second"])
        self.assertEqual([row["__sheet_row_number"] for row in rows], [2, 2])

    def test_combines_sparse_multi_row_headers(self):
        from openpyxl import Workbook

        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "multi-header.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Student", "Scores", None, "Total"])
            sheet.append([None, "Midterm", "Final", None])
            sheet.append(["An", 8, 9, 8.5])
            workbook.save(path)
            workbook.close()

            rows = materialize_spreadsheet(str(path))

        self.assertEqual(
            list(rows[0]),
            ["Student", "Scores - Midterm", "Scores - Final", "Total"],
        )
        self.assertEqual(rows[0]["Scores - Final"], 9)

    def test_dispatches_legacy_container_to_xls_reader(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "scores.xls"
            path.write_bytes(bytes.fromhex("D0CF11E0A1B11AE1") + b"payload")
            with patch(
                "data_intelligence_sdk.methods.spreadsheet._read_xls",
                return_value=[("Scores", [["Name", "Score"], ["An", 9]])],
            ) as reader:
                rows = materialize_spreadsheet(str(path))

        reader.assert_called_once_with(path)
        self.assertEqual(rows, [{"Name": "An", "Score": 9}])

    def test_registers_deterministic_materialization_capability(self):
        hub = MethodHub()
        register_spreadsheet_methods(hub)

        method = hub.get_definition("materialize_spreadsheet")

        self.assertEqual(method.trust_level, "builtin")
        self.assertIn("materialize_source", method.capability_names)
        self.assertEqual(method.metadata["output_schema"]["type"], "array")


if __name__ == "__main__":
    unittest.main()
