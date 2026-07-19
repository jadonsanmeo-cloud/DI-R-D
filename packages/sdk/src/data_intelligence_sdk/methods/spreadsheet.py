"""Deterministic Method Hub support for legacy and modern Excel workbooks."""

from __future__ import annotations

from datetime import date, datetime, time
import math
from pathlib import Path
from typing import Any, Iterable
from zipfile import BadZipFile, ZipFile, is_zipfile

from data_intelligence_sdk.runtime.method_hub import MethodHub

_SPREADSHEET_SUFFIXES = {".xls", ".xlsx", ".xlsm", ".xltx", ".xltm"}
_OLE_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")
_ZIP_SIGNATURES = (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")


class SpreadsheetReadError(ValueError):
    """Raised when a workbook cannot be identified or materialized safely."""


def _resolve_data_root(data_root: str | Path = "data") -> Path:
    root = Path(data_root)
    if root.exists():
        return root
    repo_relative = Path(__file__).resolve().parents[3] / root
    return repo_relative if repo_relative.exists() else root


def _resolve_spreadsheet(
    path: str | Path | None,
    data_root: str | Path,
) -> Path:
    root = _resolve_data_root(data_root)
    if path is not None:
        candidate = Path(path)
        if candidate.is_file():
            return candidate
        rooted = root / candidate
        if rooted.is_file():
            return rooted
        raise FileNotFoundError(f"Spreadsheet does not exist: {candidate}")

    candidates = [
        candidate
        for candidate in sorted(root.rglob("*"))
        if candidate.is_file() and candidate.suffix.lower() in _SPREADSHEET_SUFFIXES
    ]
    if not candidates:
        raise FileNotFoundError(f"No Excel workbooks were found under {root}.")
    return max(candidates, key=lambda candidate: candidate.stat().st_size)


def detect_spreadsheet_format(path: str | Path) -> str:
    """Detect ``xls`` or ``xlsx`` from the file container, not its suffix."""

    workbook_path = Path(path)
    try:
        with workbook_path.open("rb") as source:
            signature = source.read(8)
    except OSError as exc:
        raise SpreadsheetReadError(
            f"Could not inspect spreadsheet {workbook_path}: {exc}"
        ) from exc

    if signature.startswith(_OLE_SIGNATURE):
        return "xls"
    if signature.startswith(_ZIP_SIGNATURES) or is_zipfile(workbook_path):
        try:
            with ZipFile(workbook_path) as archive:
                members = set(archive.namelist())
        except (OSError, BadZipFile) as exc:
            raise SpreadsheetReadError(
                f"The spreadsheet ZIP container is invalid: {workbook_path}"
            ) from exc
        if "xl/workbook.xml" in members:
            return "xlsx"
    raise SpreadsheetReadError(
        "Unsupported spreadsheet container. Expected an Excel 97-2003 .xls "
        f"workbook or an Office Open XML .xlsx workbook: {workbook_path}"
    )


def _json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return int(value) if value.is_integer() else value
    if isinstance(value, (str, int, bool)):
        return value
    return str(value)


def _trim_row(row: Iterable[Any]) -> list[Any]:
    values = [_json_value(value) for value in row]
    while values and values[-1] is None:
        values.pop()
    return values


def _read_xlsx(path: Path) -> list[tuple[str, list[list[Any]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency contract protection
        raise SpreadsheetReadError(
            "Reading .xlsx files requires the openpyxl package."
        ) from exc

    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise SpreadsheetReadError(
            f"Could not read .xlsx workbook {path}: {exc}"
        ) from exc

    sheets: list[tuple[str, list[list[Any]]]] = []
    try:
        for worksheet in workbook.worksheets:
            rows = [_trim_row(row) for row in worksheet.iter_rows(values_only=True)]
            sheets.append((str(worksheet.title), rows))
    finally:
        workbook.close()
    return sheets


def _read_xls(path: Path) -> list[tuple[str, list[list[Any]]]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency contract protection
        raise SpreadsheetReadError(
            "Reading Excel 97-2003 .xls files requires the xlrd package."
        ) from exc

    try:
        workbook = xlrd.open_workbook(path, on_demand=True)
    except Exception as exc:
        raise SpreadsheetReadError(
            f"Could not read .xls workbook {path}: {exc}"
        ) from exc

    def cell_value(cell: Any) -> Any:
        if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
            return None
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(cell.value, workbook.datemode)
            except (OverflowError, TypeError, ValueError):
                return cell.value
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        if cell.ctype == xlrd.XL_CELL_ERROR:
            return xlrd.error_text_from_code.get(cell.value, f"#ERROR_{cell.value}")
        return cell.value

    sheets: list[tuple[str, list[list[Any]]]] = []
    try:
        for worksheet in workbook.sheets():
            rows = [
                _trim_row(cell_value(cell) for cell in worksheet.row(row_index))
                for row_index in range(worksheet.nrows)
            ]
            sheets.append((str(worksheet.name), rows))
    finally:
        workbook.release_resources()
    return sheets


def _has_values(row: list[Any]) -> bool:
    return any(value is not None and str(value).strip() for value in row)


def _detect_header_index(rows: list[list[Any]], header_row: int | None) -> int | None:
    if header_row is not None:
        if header_row < 1:
            raise ValueError("header_row must be a one-based positive integer.")
        index = header_row - 1
        if index >= len(rows):
            raise ValueError(
                f"header_row {header_row} is outside the worksheet row range."
            )
        return index

    candidates: list[tuple[tuple[int, int, float, int, int], int]] = []
    for index, row in enumerate(rows[:25]):
        positions = [
            position
            for position, value in enumerate(row)
            if value is not None and str(value).strip()
        ]
        if not positions:
            continue
        following = rows[index + 1 : index + 7]
        support = sum(
            1
            for candidate in following
            if sum(
                position < len(candidate)
                and candidate[position] is not None
                and bool(str(candidate[position]).strip())
                for position in positions
            )
            >= max(1, len(positions) // 2)
        )
        string_ratio = sum(
            isinstance(row[position], str) for position in positions
        ) / len(positions)
        distinct = len({str(row[position]).strip() for position in positions})
        score = (int(len(positions) > 1), string_ratio, distinct, support, -index)
        candidates.append((score, index))
    return max(candidates)[1] if candidates else None


def _header_span(
    rows: list[list[Any]],
    header_row: int | None,
    header_rows: int | None,
) -> tuple[int, int] | None:
    start = _detect_header_index(rows, header_row)
    if start is None:
        return None
    if header_rows is not None:
        if header_rows < 1:
            raise ValueError("header_rows must be a positive integer.")
        end = start + header_rows
        if end > len(rows):
            raise ValueError("The requested header rows exceed the worksheet range.")
        return start, end

    base_count = sum(
        value is not None and bool(str(value).strip()) for value in rows[start]
    )
    end = start + 1
    for candidate in rows[end : min(start + 3, len(rows))]:
        values = [
            value
            for value in candidate
            if value is not None and bool(str(value).strip())
        ]
        if not values:
            break
        string_ratio = sum(isinstance(value, str) for value in values) / len(values)
        if string_ratio < 0.8 or len(values) >= max(2, base_count):
            break
        end += 1
    return start, end


def _column_names(headers: list[list[Any]], width: int) -> list[str]:
    layers: list[list[str | None]] = []
    for layer_index, header in enumerate(headers):
        rendered: list[str | None] = []
        current: str | None = None
        for index in range(width):
            value = header[index] if index < len(header) else None
            label = str(value).strip() if value is not None else ""
            if label:
                current = label
                rendered.append(label)
                continue
            lower_has_label = any(
                index < len(lower)
                and lower[index] is not None
                and bool(str(lower[index]).strip())
                for lower in headers[layer_index + 1 :]
            )
            rendered.append(current if current and lower_has_label else None)
        layers.append(rendered)

    columns: list[str] = []
    counts: dict[str, int] = {}
    for index in range(width):
        parts: list[str] = []
        for layer in layers:
            label = layer[index]
            if label and label.casefold() not in {item.casefold() for item in parts}:
                parts.append(label)
        base = " - ".join(parts)
        base = base or f"column_{index + 1}"
        counts[base] = counts.get(base, 0) + 1
        columns.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return columns


def _materialize_sheet(
    sheet_name: str,
    rows: list[list[Any]],
    *,
    header_row: int | None,
    header_rows: int | None,
    include_sheet_metadata: bool,
) -> list[dict[str, Any]]:
    span = _header_span(rows, header_row, header_rows)
    if span is None:
        return []
    header_index, data_start = span
    data_rows = [row for row in rows[data_start:] if _has_values(row)]
    if not data_rows:
        return []
    width = max(
        [
            *(len(row) for row in rows[header_index:data_start]),
            *(len(row) for row in data_rows),
        ]
    )
    columns = _column_names(rows[header_index:data_start], width)
    materialized = []
    for offset, row in enumerate(rows[data_start:], start=data_start + 1):
        if not _has_values(row):
            continue
        record = {
            column: _json_value(row[index]) if index < len(row) else None
            for index, column in enumerate(columns)
        }
        if include_sheet_metadata:
            record["__sheet_name"] = sheet_name
            record["__sheet_row_number"] = offset
        materialized.append(record)
    return materialized


def _select_sheets(
    sheets: list[tuple[str, list[list[Any]]]],
    sheet_name: str | int | None,
    include_all_sheets: bool,
) -> list[tuple[str, list[list[Any]]]]:
    non_empty = [item for item in sheets if any(_has_values(row) for row in item[1])]
    if sheet_name is not None and include_all_sheets:
        raise ValueError("sheet_name and include_all_sheets cannot be used together.")
    if sheet_name is None:
        return non_empty if include_all_sheets else non_empty[:1]
    if isinstance(sheet_name, int) or str(sheet_name).strip().isdigit():
        index = int(sheet_name)
        if index < 0 or index >= len(sheets):
            raise ValueError(f"Worksheet index {index} is outside the workbook range.")
        return [sheets[index]]
    requested = str(sheet_name).strip().casefold()
    selected = [item for item in sheets if item[0].casefold() == requested]
    if not selected:
        available = ", ".join(name for name, _ in sheets)
        raise ValueError(
            f"Worksheet {sheet_name!r} was not found. Available: {available}"
        )
    return selected


def materialize_spreadsheet(
    path: str | None = None,
    data_root: str = "data",
    sheet_name: str | int | None = None,
    header_row: int | None = None,
    header_rows: int | None = None,
    include_all_sheets: bool = False,
    max_rows: int | None = None,
) -> list[dict[str, Any]]:
    """Read an XLS/XLSX workbook into JSON-safe row dictionaries.

    The workbook container is detected from its bytes. By default the first
    non-empty worksheet is selected and a header row is detected from the first
    25 rows. Callers can select a worksheet, force a one-based header row and
    header depth, or combine all non-empty worksheets with ``__sheet_*`` lineage
    fields.
    """

    workbook_path = _resolve_spreadsheet(path, data_root)
    workbook_format = detect_spreadsheet_format(workbook_path)
    sheets = (
        _read_xls(workbook_path)
        if workbook_format == "xls"
        else _read_xlsx(workbook_path)
    )
    selected = _select_sheets(sheets, sheet_name, include_all_sheets)
    if max_rows is not None and int(max_rows) < 1:
        raise ValueError("max_rows must be a positive integer when provided.")

    output: list[dict[str, Any]] = []
    include_metadata = include_all_sheets and len(selected) > 1
    for selected_name, rows in selected:
        output.extend(
            _materialize_sheet(
                selected_name,
                rows,
                header_row=header_row,
                header_rows=header_rows,
                include_sheet_metadata=include_metadata,
            )
        )
        if max_rows is not None and len(output) >= int(max_rows):
            return output[: int(max_rows)]
    return output


def register_spreadsheet_methods(method_hub: MethodHub) -> None:
    """Register deterministic spreadsheet materialization in a Method Hub."""

    description = (
        "Materialize Excel 97-2003 XLS and modern XLSX/XLSM workbooks into "
        "JSON-safe table rows with optional sheet and header selection."
    )
    method_hub.register(
        "materialize_spreadsheet",
        materialize_spreadsheet,
        capability_names=[
            "materialize_source",
            "load_spreadsheet",
            "read_excel",
            "inspect_tabular_data",
            "generate_report",
        ],
        trust_level="builtin",
        metadata={
            "category": "spreadsheet",
            "deterministic": True,
            "side_effects": False,
            "description": description,
            "parameters_schema": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Workbook path."},
                    "data_root": {
                        "type": "string",
                        "description": "Fallback data root.",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Worksheet name or zero-based index.",
                    },
                    "header_row": {
                        "type": "integer",
                        "description": (
                            "One-based header row; auto-detected when omitted."
                        ),
                    },
                    "header_rows": {
                        "type": "integer",
                        "description": (
                            "Number of header rows to combine; auto-detected "
                            "when omitted."
                        ),
                    },
                    "include_all_sheets": {
                        "type": "boolean",
                        "description": "Combine all non-empty worksheets.",
                    },
                    "max_rows": {
                        "type": "integer",
                        "description": "Optional positive global row limit.",
                    },
                },
                "required": [],
            },
            "output_schema": {"type": "array", "items": {"type": "object"}},
            "use_when": [
                "A plan step must load or materialize an XLS/XLSX workbook.",
                "A report needs structured rows from a local Excel source.",
            ],
        },
        version="1.0.0",
        description=description,
        tags=["spreadsheet", "excel", "xls", "xlsx", "materialization"],
        status="stable",
        priority=155,
        source=__name__,
    )
